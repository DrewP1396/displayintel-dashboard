"""
Factories Page - Display Intelligence Dashboard
Factory database, utilization tracking, and capacity analysis.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.styling import get_css, get_plotly_theme, apply_chart_theme, format_with_commas, format_percent
from utils.database import DatabaseManager
from utils.exports import create_download_buttons

# Page config
st.set_page_config(
    page_title="Factories - Display Intelligence",
    page_icon="🏭",
    layout="wide"
)

# Apply styling
st.markdown(get_css(), unsafe_allow_html=True)

# Check authentication (restore session from cookie if available)

if not st.session_state.get("password_correct", False):
    st.warning("Please login from the main page.")
    st.stop()


def natural_sort_key(s):
    """Sort strings with embedded numbers naturally (B1, B2, B10 not B1, B10, B2)."""
    if pd.isna(s):
        return []
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]


# Page header
st.markdown("""
    <h1>🏭 Factory Intelligence</h1>
    <p style="color: #86868B; font-size: 1.1rem; margin-bottom: 2rem;">
        Comprehensive database of display manufacturing facilities worldwide
    </p>
""", unsafe_allow_html=True)

# Filters in sidebar
with st.sidebar:
    st.markdown("### Filters")

    manufacturer = st.selectbox(
        "Manufacturer",
        options=DatabaseManager.get_manufacturers(),
        key="factory_manufacturer"
    )

    # Factory-specific filter (depends on manufacturer selection)
    factory_options = DatabaseManager.get_factory_names(manufacturer)
    selected_factory = st.selectbox(
        "Factory",
        options=factory_options,
        key="factory_specific"
    )

    st.divider()

    # Only show these filters in "All Factories" view
    if selected_factory == "All Factories":
        technology = st.selectbox(
            "Technology",
            options=DatabaseManager.get_technologies(),
            key="factory_technology"
        )

        region = st.selectbox(
            "Region",
            options=DatabaseManager.get_regions(),
            key="factory_region"
        )

        status_options = ["All", "operating", "constructing", "planned", "closed"]
        status = st.selectbox(
            "Status",
            options=status_options,
            key="factory_status"
        )

        st.divider()
    else:
        # Set defaults when viewing specific factory
        technology = "All"
        region = "All"
        status = "All"

    # Date range for utilization
    st.markdown("### Utilization Date Range")
    min_date, max_date = DatabaseManager.get_date_range()

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "Start",
            value=date(2023, 1, 1),
            min_value=datetime.strptime(min_date, "%Y-%m-%d").date(),
            max_value=datetime.strptime(max_date, "%Y-%m-%d").date(),
            key="util_start"
        )
    with col2:
        # Default end date to today (not max DB date which may be years in the future)
        default_end = min(date.today(), datetime.strptime(max_date, "%Y-%m-%d").date())
        end_date = st.date_input(
            "End",
            value=default_end,
            min_value=datetime.strptime(min_date, "%Y-%m-%d").date(),
            max_value=datetime.strptime(max_date, "%Y-%m-%d").date(),
            key="util_end"
        )

# Get theme colors
theme = get_plotly_theme()
colors = theme['color_discrete_sequence']

# =============================================================================
# Factory Detail View (when specific factory is selected)
# =============================================================================
if selected_factory != "All Factories":
    # Get factory data (may have multiple entries for different backplanes)
    factory_df = DatabaseManager.get_factory_by_name(selected_factory)

    if factory_df is not None and len(factory_df) > 0:
        # Use first entry for general info
        factory = factory_df.iloc[0]

        # Get all backplane types for this factory
        backplanes = factory_df['backplane'].dropna().unique().tolist()
        backplane_str = ", ".join(backplanes) if backplanes else "-"

        # Get MP Ramp dates by backplane from factories table
        def format_quarter(date_val):
            """Format date as quarter (e.g., CQ1'19)."""
            if pd.isna(date_val) or not date_val:
                return None
            try:
                if isinstance(date_val, str):
                    if len(date_val) == 4:  # Just year like "2015"
                        return f"CQ1'{date_val[2:]}"
                    date_val = pd.to_datetime(date_val)
                q = (date_val.month - 1) // 3 + 1
                return f"CQ{q}'{str(date_val.year)[2:]}"
            except:
                return str(date_val)[:10] if date_val else None

        ramp_by_bp = {}
        for _, row in factory_df.iterrows():
            bp = row.get('backplane', 'Unknown')
            mp_ramp = row.get('mp_ramp_date')
            if mp_ramp and bp:
                ramp_by_bp[bp] = format_quarter(mp_ramp)

        # Get earliest ramp date for display
        ramp_dates = [r for r in ramp_by_bp.values() if r]
        ramp_date = min(ramp_dates) if ramp_dates else None

        # Header with back context
        tech = factory.get('technology', 'Unknown') or 'Unknown'
        gen = factory.get('generation', '') or ''
        st.markdown(f"""
            <div style="margin-bottom: 1rem;">
                <span style="color: #86868B; font-size: 0.9rem;">
                    {factory.get('manufacturer', 'Unknown')} &gt; {selected_factory}
                </span>
                <span style="background: {'#007AFF' if tech == 'OLED' else '#34C759'}; color: white; padding: 2px 8px; border-radius: 4px; font-size: 0.75rem; margin-left: 8px;">
                    {tech} {gen}
                </span>
            </div>
        """, unsafe_allow_html=True)

        # Factory Summary Card
        st.markdown("### Factory Summary")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Factory", selected_factory)
            st.metric("Generation", factory.get('generation', '-') or "-")

        with col2:
            st.metric("Location", factory.get('location', '-') or "-")
            st.metric("Technology", factory.get('technology', '-') or "-")

        with col3:
            st.metric("Region", factory.get('region', '-') or "-")
            st.metric("Backplanes", backplane_str)

        with col4:
            st.metric("Status", (factory.get('status', '-') or "-").title())
            st.metric("First Ramp", ramp_date or "Not yet")

        st.divider()

        # Capacity by Backplane Technology
        st.markdown("### Capacity by Backplane Technology")

        # Get capacity data for this factory and related production lines
        backplane_df = DatabaseManager.get_capacity_by_backplane(
            manufacturer=factory.get('manufacturer'),
            factory_name=selected_factory
        )

        if len(backplane_df) > 0:
            # Group by backplane and sum capacity
            bp_summary = backplane_df.groupby('backplane').agg({
                'capacity_ksheets': 'sum',
                'actual_input_ksheets': 'sum',
                'factory_name': 'count'
            }).reset_index()
            bp_summary.columns = ['Backplane', 'Capacity (K/mo)', 'Input (K/mo)', 'Lines']

            # Calculate total
            total_capacity = bp_summary['Capacity (K/mo)'].sum()
            total_input = bp_summary['Input (K/mo)'].sum()

            # Display metrics with ramp dates
            cols = st.columns(len(bp_summary) + 1)

            # Total column
            with cols[0]:
                st.metric("Total Capacity", f"{total_capacity:,.1f}K/mo")

            # Per-backplane columns with ramp date
            for i, row in bp_summary.iterrows():
                with cols[i + 1]:
                    bp_name = row['Backplane'] or 'Unknown'
                    bp_ramp = ramp_by_bp.get(bp_name, '')
                    label = f"{bp_name}" + (f" ({bp_ramp})" if bp_ramp else "")
                    st.metric(label, f"{row['Capacity (K/mo)']:,.1f}K/mo")

            # Show detailed breakdown table
            with st.expander("View Production Lines"):
                bp_display = backplane_df[['factory_name', 'backplane', 'generation', 'capacity_ksheets', 'actual_input_ksheets', 'utilization_pct']].copy()
                bp_display['capacity_ksheets'] = bp_display['capacity_ksheets'].apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "-")
                bp_display['actual_input_ksheets'] = bp_display['actual_input_ksheets'].apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "-")
                bp_display['utilization_pct'] = bp_display['utilization_pct'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")

                st.dataframe(
                    bp_display,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "factory_name": st.column_config.TextColumn("Line", width="medium"),
                        "backplane": st.column_config.TextColumn("Backplane", width="small"),
                        "generation": st.column_config.TextColumn("Gen", width="small"),
                        "capacity_ksheets": st.column_config.TextColumn("Capacity (K/mo)", width="small"),
                        "actual_input_ksheets": st.column_config.TextColumn("Input (K/mo)", width="small"),
                        "utilization_pct": st.column_config.TextColumn("Util %", width="small")
                    }
                )
        else:
            st.info("No capacity data available.")

        # Capacity Installments by Quarter
        st.markdown("### Capacity Installments by Quarter")

        # Get full utilization history to show capacity additions
        full_util_df = DatabaseManager.get_utilization(factory_name=selected_factory)

        if len(full_util_df) > 0:
            # Group by quarter and backplane, get capacity
            full_util_df['quarter'] = pd.to_datetime(full_util_df['date']).dt.to_period('Q').astype(str)

            # Get capacity per quarter per backplane
            quarterly_cap = full_util_df.groupby(['quarter', 'backplane']).agg({
                'capacity_ksheets': 'max'  # Use max capacity for each quarter
            }).reset_index()

            # Calculate capacity additions (difference from previous quarter)
            quarterly_cap = quarterly_cap.sort_values(['backplane', 'quarter'])
            quarterly_cap['capacity_change'] = quarterly_cap.groupby('backplane')['capacity_ksheets'].diff().fillna(quarterly_cap['capacity_ksheets'])

            # Only show quarters with capacity additions
            additions = quarterly_cap[quarterly_cap['capacity_change'] > 0.5].copy()

            if len(additions) > 0:
                # Sort additions chronologically
                additions = additions.sort_values('quarter')

                # Get all quarters in sorted order for x-axis
                all_quarters = sorted(additions['quarter'].unique())

                # Create stacked bar chart
                fig = go.Figure()

                for bp in additions['backplane'].unique():
                    bp_data = additions[additions['backplane'] == bp]
                    fig.add_trace(go.Bar(
                        x=bp_data['quarter'].tolist(),
                        y=bp_data['capacity_change'].tolist(),
                        name=bp,
                        hovertemplate=f'{bp}<br>%{{x}}<br>+%{{y:,.1f}}K/mo<extra></extra>'
                    ))

                apply_chart_theme(fig)
                fig.update_layout(
                    xaxis_title="Quarter",
                    yaxis_title="Capacity Added (K/mo)",
                    height=300,
                    barmode='group',
                    showlegend=True,
                    xaxis={'categoryorder': 'array', 'categoryarray': all_quarters}
                )
                st.plotly_chart(fig, use_container_width=True)

                # Show table of additions
                with st.expander("View Capacity Addition Details"):
                    additions_display = additions[['quarter', 'backplane', 'capacity_change', 'capacity_ksheets']].copy()
                    additions_display = additions_display.sort_values('quarter')
                    additions_display.columns = ['Quarter', 'Backplane', 'Added (K/mo)', 'Total (K/mo)']
                    additions_display['Added (K/mo)'] = additions_display['Added (K/mo)'].apply(lambda x: f"{x:,.1f}")
                    additions_display['Total (K/mo)'] = additions_display['Total (K/mo)'].apply(lambda x: f"{x:,.1f}")
                    st.dataframe(additions_display, use_container_width=True, hide_index=True)
            else:
                st.info("No capacity additions found in the data.")
        else:
            st.info("No capacity history available.")

        st.divider()

        # Get utilization data for this factory (all backplane variants)
        util_df = DatabaseManager.get_utilization(
            start_date=start_date.strftime("%Y-%m-%d"),
            end_date=end_date.strftime("%Y-%m-%d"),
            factory_name=selected_factory
        )

        # Filter out future quarters with no actual data (all 0%)
        if len(util_df) > 0:
            current_quarter_end = pd.Timestamp.today().to_period('Q').end_time.strftime("%Y-%m-%d")
            util_df = util_df[util_df['date'] <= current_quarter_end]

        # Current metrics from latest utilization with actual data (not projections)
        if len(util_df) > 0:
            # Get latest date with actual input > 0
            util_with_data = util_df[util_df['actual_input_ksheets'] > 0]
            if len(util_with_data) > 0:
                latest_date = util_with_data['date'].max()
            else:
                latest_date = util_df['date'].max()
            latest_data = util_df[util_df['date'] == latest_date]

            total_capacity = latest_data['capacity_ksheets'].sum()
            total_input = latest_data['actual_input_ksheets'].sum()
            avg_util = (total_input / total_capacity * 100) if total_capacity > 0 else 0

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Current Utilization", f"{avg_util:.1f}%")
            with col2:
                st.metric("Total Capacity (K/mo)", f"{total_capacity:,.1f}")
            with col3:
                st.metric("Total Input (K/mo)", f"{total_input:,.1f}")

            st.divider()

            # Utilization Timeline by Backplane
            st.markdown("### Utilization Timeline by Backplane")

            fig = go.Figure()

            # Group by date and backplane
            backplanes_in_data = sorted(util_df['backplane'].dropna().unique())

            for i, bp in enumerate(backplanes_in_data):
                bp_data = util_df[util_df['backplane'] == bp].groupby('date').agg({
                    'capacity_ksheets': 'sum',
                    'actual_input_ksheets': 'sum'
                }).reset_index()
                bp_data['utilization_pct'] = (bp_data['actual_input_ksheets'] / bp_data['capacity_ksheets'] * 100)

                fig.add_trace(go.Scatter(
                    x=bp_data['date'].tolist(),
                    y=bp_data['utilization_pct'].tolist(),
                    mode='lines+markers',
                    name=f'{bp} ({bp_data["capacity_ksheets"].iloc[-1]:,.0f}K cap)',
                    line=dict(color=colors[i % len(colors)], width=2),
                    marker=dict(size=6),
                    hovertemplate=f'{bp}<br>%{{x}}<br>Utilization: %{{y:.1f}}%<br>Capacity: {bp_data["capacity_ksheets"].iloc[-1]:,.1f}K/mo<extra></extra>'
                ))

            # Note if utilization rates are identical across backplanes
            if len(backplanes_in_data) > 1:
                # Check if util % is the same across backplanes on shared dates
                shared_dates = util_df.groupby('date').filter(lambda x: x['backplane'].nunique() > 1)
                if len(shared_dates) > 0:
                    util_spread = shared_dates.groupby('date')['utilization_pct'].std().mean()
                    if util_spread < 0.1:
                        st.caption("Note: Source data applies the same utilization rate across backplane lines. "
                                   "Lines overlap on the chart. See Capacity vs Input below for per-backplane breakdown.")

            # Add ramp date annotation if available
            if ramp_date and ramp_date in util_df['date'].values:
                fig.add_vline(
                    x=ramp_date,
                    line_dash="dash",
                    line_color="green",
                    annotation_text="First Ramp",
                    annotation_position="top"
                )

            apply_chart_theme(fig)
            fig.update_layout(
                xaxis_title="Date",
                yaxis_title="Utilization (%)",
                height=400,
                hovermode='x unified'
            )

            st.plotly_chart(fig, use_container_width=True)

            # Capacity and Input Chart - total factory snapshot
            st.markdown("### Monthly Capacity vs Actual Input")

            # Aggregate across all backplanes for factory-level view
            factory_totals = util_df.groupby('date').agg({
                'capacity_ksheets': 'sum',
                'actual_input_ksheets': 'sum'
            }).reset_index()
            factory_totals['utilization_pct'] = (
                factory_totals['actual_input_ksheets'] / factory_totals['capacity_ksheets'] * 100
            ).round(1)

            fig2 = go.Figure()

            fig2.add_trace(go.Bar(
                x=factory_totals['date'].tolist(),
                y=factory_totals['capacity_ksheets'].tolist(),
                name='Total Capacity',
                marker_color=colors[0],
                opacity=0.7,
                hovertemplate='Capacity: %{y:,.1f}K/mo<extra></extra>'
            ))

            fig2.add_trace(go.Scatter(
                x=factory_totals['date'].tolist(),
                y=factory_totals['actual_input_ksheets'].tolist(),
                mode='lines+markers',
                name='Actual Input',
                line=dict(color=colors[1], width=2),
                hovertemplate='Input: %{y:,.1f}K/mo<extra></extra>'
            ))

            apply_chart_theme(fig2)
            fig2.update_layout(
                xaxis_title="Date",
                yaxis_title="K Sheets / Month",
                height=350,
                barmode='overlay',
                hovermode='x unified'
            )

            st.plotly_chart(fig2, use_container_width=True)

        else:
            st.info("No utilization data available for this factory in the selected date range.")

        st.divider()

        # Equipment Orders for this factory (get orders for all backplane variants)
        st.markdown("### Equipment Orders")

        # Get all factory_ids for this factory name
        factory_ids = factory_df['factory_id'].tolist()
        equip_dfs = [DatabaseManager.get_equipment_orders_for_factory(fid) for fid in factory_ids]
        equip_df = pd.concat(equip_dfs, ignore_index=True) if equip_dfs else pd.DataFrame()

        if len(equip_df) > 0:
            # Summary metrics
            total_orders = len(equip_df)
            total_value = equip_df['amount_usd'].sum() if 'amount_usd' in equip_df.columns else 0
            total_units = equip_df['units'].sum() if 'units' in equip_df.columns else 0

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Orders", format_with_commas(total_orders))
            with col2:
                if total_value and total_value > 0:
                    st.metric("Total Value", f"${total_value/1e6:.1f}M" if total_value < 1e9 else f"${total_value/1e9:.2f}B")
                else:
                    st.metric("Total Value", "-")
            with col3:
                st.metric("Total Units", format_with_commas(total_units) if total_units else "-")

            # Equipment orders table - format for display
            display_cols = ['po_year', 'po_quarter', 'vendor', 'equipment_type', 'units', 'amount_usd']
            available_cols = [c for c in display_cols if c in equip_df.columns]

            equip_display = equip_df[available_cols].head(100).copy()

            # Format columns for cleaner display
            if 'units' in equip_display.columns:
                equip_display['units'] = equip_display['units'].apply(lambda x: f"{int(x):,}" if pd.notna(x) and x > 0 else "-")
            if 'amount_usd' in equip_display.columns:
                equip_display['amount_usd'] = equip_display['amount_usd'].apply(
                    lambda x: f"${x:,.0f}" if pd.notna(x) and x > 0 else "-"
                )

            st.dataframe(
                equip_display,
                use_container_width=True,
                hide_index=True,
                height=300,
                column_config={
                    "po_year": st.column_config.TextColumn("Year", width="small"),
                    "po_quarter": st.column_config.TextColumn("Qtr", width="small"),
                    "vendor": st.column_config.TextColumn("Vendor", width="medium"),
                    "equipment_type": st.column_config.TextColumn("Equipment", width="medium"),
                    "units": st.column_config.TextColumn("Units", width="small"),
                    "amount_usd": st.column_config.TextColumn("Value", width="medium")
                }
            )
        else:
            st.info("No equipment orders found for this factory.")

    else:
        st.error(f"Factory '{selected_factory}' not found.")

else:
    # =============================================================================
    # All Factories View (original tabs)
    # =============================================================================

    # Main content tabs
    tab1, tab2, tab3, tab4 = st.tabs(["Factory Database", "Utilization Analysis", "Capacity Overview", "Factory Comparison"])

    # Tab 1: Factory Database
    with tab1:
        # Load factory data
        try:
            factories_df = DatabaseManager.get_factories(
                manufacturer=manufacturer,
                technology=technology,
                region=region,
                status=status
            )
        except Exception as e:
            st.error(f"Error loading factory data: {str(e)}")
            st.stop()

        # Add ramp dates to factories
        ramp_dates = DatabaseManager.get_all_factory_ramp_dates()
        factories_df['ramp_date'] = factories_df['factory_id'].map(ramp_dates)

        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Total Factories", format_with_commas(len(factories_df)))

        with col2:
            operating = len(factories_df[factories_df['status'] == 'operating'])
            st.metric("Operating", format_with_commas(operating))

        with col3:
            unique_mfrs = factories_df['manufacturer'].nunique()
            st.metric("Manufacturers", format_with_commas(unique_mfrs))

        with col4:
            unique_regions = factories_df['region'].nunique()
            st.metric("Regions", format_with_commas(unique_regions))

        st.divider()

        # Charts row
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Factories by Manufacturer")
            # Filter out NULL/empty values and get counts
            mfr_data = factories_df[factories_df['manufacturer'].notna() & (factories_df['manufacturer'] != '')]
            mfr_counts = mfr_data['manufacturer'].value_counts().head(15)

            if len(mfr_counts) > 0:
                fig = px.bar(
                    x=mfr_counts.values.tolist(),
                    y=mfr_counts.index.tolist(),
                    orientation='h',
                    labels={'x': 'Number of Factories', 'y': 'Manufacturer'}
                )
                fig.update_traces(marker_color=colors[0], hovertemplate='%{y}: %{x} factories<extra></extra>')
                apply_chart_theme(fig)
                fig.update_layout(
                    showlegend=False,
                    xaxis_title="Number of Factories",
                    yaxis_title="",
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No manufacturer data available")

        with col2:
            st.markdown("#### Factories by Technology")
            # Filter out NULL/empty values
            tech_data = factories_df[factories_df['technology'].notna() & (factories_df['technology'] != '')]
            tech_counts = tech_data['technology'].value_counts()

            if len(tech_counts) > 0:
                fig = px.pie(
                    values=tech_counts.values.tolist(),
                    names=tech_counts.index.tolist(),
                    color_discrete_sequence=colors,
                    hole=0.4
                )
                fig.update_traces(hovertemplate='%{label}: %{value} factories (%{percent})<extra></extra>')
                apply_chart_theme(fig)
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No technology data available")

        st.divider()

        # Data table
        st.markdown("#### Factory Database")

        # Select display columns with ramp_date
        display_cols = [
            'factory_id', 'manufacturer', 'factory_name', 'location', 'region',
            'technology', 'generation', 'application_category', 'status', 'ramp_date'
        ]
        available_cols = [c for c in display_cols if c in factories_df.columns]

        # Sort naturally by factory_name
        display_df = factories_df[available_cols].copy()
        if 'factory_name' in display_df.columns:
            display_df = display_df.iloc[display_df['factory_name'].map(natural_sort_key).argsort()]

        column_config = {
            "factory_id": st.column_config.TextColumn("Factory ID", width="medium"),
            "manufacturer": st.column_config.TextColumn("Manufacturer", width="small"),
            "factory_name": st.column_config.TextColumn("Factory Name", width="medium"),
            "location": st.column_config.TextColumn("Location", width="medium"),
            "region": st.column_config.TextColumn("Region", width="small"),
            "technology": st.column_config.TextColumn("Tech", width="small"),
            "generation": st.column_config.TextColumn("Gen", width="small"),
            "application_category": st.column_config.TextColumn("Application", width="small"),
            "status": st.column_config.TextColumn("Status", width="small"),
            "ramp_date": st.column_config.TextColumn("Ramp Date", width="small")
        }

        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            height=400,
            column_config=column_config
        )

        # Export buttons
        st.markdown("<br>", unsafe_allow_html=True)
        create_download_buttons(factories_df, "factories", "Factory Database Report")


    # Tab 2: Utilization Analysis
    with tab2:
        # Load utilization data
        try:
            util_df = DatabaseManager.get_utilization(
                start_date=start_date.strftime("%Y-%m-%d"),
                end_date=end_date.strftime("%Y-%m-%d"),
                manufacturer=manufacturer if manufacturer != "All" else None
            )
        except Exception as e:
            st.error(f"Error loading utilization data: {str(e)}")
            st.stop()

        # Filter out future quarters with no actual data
        if len(util_df) > 0:
            current_quarter_end = pd.Timestamp.today().to_period('Q').end_time.strftime("%Y-%m-%d")
            util_df = util_df[util_df['date'] <= current_quarter_end]

        if len(util_df) > 0:
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                avg_util = util_df['utilization_pct'].mean()
                st.metric("Avg Utilization", format_percent(avg_util))

            with col2:
                max_util = util_df['utilization_pct'].max()
                st.metric("Max Utilization", format_percent(max_util))

            with col3:
                total_capacity = util_df.groupby('date')['capacity_ksheets'].sum().mean()
                st.metric("Avg Monthly Capacity", f"{total_capacity:,.0f}K/mo")

            with col4:
                factories_count = util_df['factory_id'].nunique()
                st.metric("Factories Tracked", format_with_commas(factories_count))

            st.divider()

            # Utilization over time
            st.markdown("#### Utilization Trends Over Time")

            util_by_date = util_df.groupby('date').agg({
                'utilization_pct': 'mean',
                'capacity_ksheets': 'sum',
                'actual_input_ksheets': 'sum'
            }).reset_index()

            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=util_by_date['date'].tolist(),
                y=util_by_date['utilization_pct'].tolist(),
                mode='lines+markers',
                name='Utilization %',
                line=dict(color=colors[0], width=2),
                marker=dict(size=6),
                hovertemplate='%{x}<br>Utilization: %{y:.1f}%<extra></extra>'
            ))

            apply_chart_theme(fig)
            fig.update_layout(
                xaxis_title="Date",
                yaxis_title="Utilization (%)",
                height=400,
                hovermode='x unified'
            )

            st.plotly_chart(fig, use_container_width=True)

            # Utilization by manufacturer
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Utilization by Manufacturer")

                # Filter out NULL/empty manufacturers
                util_mfr_df = util_df[util_df['manufacturer'].notna() & (util_df['manufacturer'] != '')]
                util_by_mfr = util_mfr_df.groupby('manufacturer')['utilization_pct'].mean().sort_values(ascending=True)

                if len(util_by_mfr) > 0:
                    fig = px.bar(
                        x=util_by_mfr.values.tolist(),
                        y=util_by_mfr.index.tolist(),
                        orientation='h'
                    )
                    fig.update_traces(marker_color=colors[0], hovertemplate='%{y}: %{x:.1f}%<extra></extra>')
                    apply_chart_theme(fig)
                    fig.update_layout(
                        showlegend=False,
                        xaxis_title="Average Utilization (%)",
                        yaxis_title="",
                        height=400
                    )
                    st.plotly_chart(fig, use_container_width=True)

            with col2:
                st.markdown("#### Utilization Distribution")

                fig = px.histogram(
                    util_df,
                    x='utilization_pct',
                    nbins=30
                )
                fig.update_traces(marker_color=colors[0])
                apply_chart_theme(fig)
                fig.update_layout(
                    showlegend=False,
                    xaxis_title="Utilization (%)",
                    yaxis_title="Count",
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)

            st.divider()

            # Detailed utilization table
            st.markdown("#### Monthly Utilization Details")

            util_display_cols = [
                'date', 'manufacturer', 'factory_name', 'technology', 'utilization_pct',
                'capacity_ksheets', 'actual_input_ksheets'
            ]
            available_cols = [c for c in util_display_cols if c in util_df.columns]

            # Format for display
            util_display = util_df[available_cols].head(500).copy()
            if 'utilization_pct' in util_display.columns:
                util_display['utilization_pct'] = util_display['utilization_pct'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")
            if 'capacity_ksheets' in util_display.columns:
                util_display['capacity_ksheets'] = util_display['capacity_ksheets'].apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "-")
            if 'actual_input_ksheets' in util_display.columns:
                util_display['actual_input_ksheets'] = util_display['actual_input_ksheets'].apply(lambda x: f"{x:,.1f}" if pd.notna(x) else "-")

            st.dataframe(
                util_display,
                use_container_width=True,
                hide_index=True,
                height=400,
                column_config={
                    "date": st.column_config.TextColumn("Date", width="small"),
                    "manufacturer": st.column_config.TextColumn("Mfr", width="small"),
                    "factory_name": st.column_config.TextColumn("Factory", width="medium"),
                    "technology": st.column_config.TextColumn("Tech", width="small"),
                    "utilization_pct": st.column_config.TextColumn("Util %", width="small"),
                    "capacity_ksheets": st.column_config.TextColumn("Capacity (K/mo)", width="small"),
                    "actual_input_ksheets": st.column_config.TextColumn("Input (K/mo)", width="small")
                }
            )

            create_download_buttons(util_df, "utilization", "Utilization Report")

        else:
            st.info("No utilization data available for the selected filters.")


    # Tab 3: Capacity Overview
    with tab3:
        # Load utilization data for capacity analysis
        try:
            util_df = DatabaseManager.get_utilization(
                start_date=start_date.strftime("%Y-%m-%d"),
                end_date=end_date.strftime("%Y-%m-%d")
            )
        except Exception as e:
            st.error(f"Error loading capacity data: {str(e)}")
            st.stop()

        # Filter out future quarters with no actual data
        if len(util_df) > 0:
            current_quarter_end = pd.Timestamp.today().to_period('Q').end_time.strftime("%Y-%m-%d")
            util_df = util_df[util_df['date'] <= current_quarter_end]

        if len(util_df) > 0:
            st.markdown("#### Total Industry Capacity Over Time")

            capacity_by_date = util_df.groupby('date').agg({
                'capacity_ksheets': 'sum',
                'actual_input_ksheets': 'sum',
                'capacity_sqm_k': 'sum'
            }).reset_index()

            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=capacity_by_date['date'].tolist(),
                y=capacity_by_date['capacity_ksheets'].tolist(),
                mode='lines',
                name='Total Capacity',
                fill='tozeroy',
                line=dict(color=colors[0], width=2),
                fillcolor='rgba(0, 122, 255, 0.1)',
                hovertemplate='Capacity: %{y:,.0f}K<extra></extra>'
            ))

            fig.add_trace(go.Scatter(
                x=capacity_by_date['date'].tolist(),
                y=capacity_by_date['actual_input_ksheets'].tolist(),
                mode='lines',
                name='Actual Input',
                line=dict(color=colors[1], width=2),
                hovertemplate='Input: %{y:,.0f}K<extra></extra>'
            ))

            apply_chart_theme(fig)
            fig.update_layout(
                xaxis_title="Date",
                yaxis_title="K Sheets",
                height=400,
                hovermode='x unified'
            )

            st.plotly_chart(fig, use_container_width=True)

            st.divider()

            # Capacity by manufacturer
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Capacity Share by Manufacturer")

                latest_date = util_df['date'].max()
                # Filter out NULL/empty manufacturers
                capacity_df = util_df[(util_df['date'] == latest_date) &
                                       util_df['manufacturer'].notna() &
                                       (util_df['manufacturer'] != '')]
                latest_capacity = capacity_df.groupby('manufacturer')['capacity_ksheets'].sum()

                if len(latest_capacity) > 0:
                    fig = px.pie(
                        values=latest_capacity.values.tolist(),
                        names=latest_capacity.index.tolist(),
                        color_discrete_sequence=colors,
                        hole=0.4
                    )
                    fig.update_traces(hovertemplate='%{label}: %{value:,.0f}K (%{percent})<extra></extra>')
                    apply_chart_theme(fig)
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)

            with col2:
                st.markdown("#### Capacity by Technology")

                # Filter out NULL/empty technologies
                tech_util_df = util_df[util_df['technology'].notna() & (util_df['technology'] != '')]
                capacity_by_tech = tech_util_df.groupby('technology')['capacity_ksheets'].sum()

                if len(capacity_by_tech) > 0:
                    fig = px.bar(
                        x=capacity_by_tech.index.tolist(),
                        y=capacity_by_tech.values.tolist()
                    )
                    fig.update_traces(marker_color=colors[0], hovertemplate='%{x}: %{y:,.0f}K<extra></extra>')
                    apply_chart_theme(fig)
                    fig.update_layout(
                        showlegend=False,
                        xaxis_title="Technology",
                        yaxis_title="Total Capacity (K Sheets)",
                        height=400
                    )
                    st.plotly_chart(fig, use_container_width=True)

            # Regional capacity breakdown
            st.divider()
            st.markdown("#### Regional Capacity Distribution")

            # Filter out NULL/empty regions
            region_df = util_df[util_df['region'].notna() & (util_df['region'] != '')]
            capacity_by_region = region_df.groupby('region')['capacity_ksheets'].sum().sort_values(ascending=False)

            if len(capacity_by_region) > 0:
                fig = px.bar(
                    x=capacity_by_region.index.tolist(),
                    y=capacity_by_region.values.tolist()
                )
                fig.update_traces(marker_color=colors[0], hovertemplate='%{x}: %{y:,.0f}K<extra></extra>')
                apply_chart_theme(fig)
                fig.update_layout(
                    showlegend=False,
                    xaxis_title="Region",
                    yaxis_title="Total Capacity (K Sheets)",
                    height=350
                )
                st.plotly_chart(fig, use_container_width=True)

        else:
            st.info("No capacity data available for the selected date range.")

    # Tab 4: Factory Comparison
    with tab4:
        # ── Constants ──
        _DEP_YEARS = {"S Korea": 5, "China": 7, "Taiwan": 7, "Japan": 5,
                       "Singapore": 5, "India": 7}

        # TODO: Add investment data from historical CapSpendReport files
        # Structure: {(manufacturer, factory1, phase): amount_usd}
        _PHASE_INVESTMENT: dict = {}

        # ── Helper functions ──
        def _fmt_mp(val):
            """Format MP ramp date for display."""
            if not val or (isinstance(val, float) and pd.isna(val)):
                return "-"
            s = str(val).strip()
            if s.startswith("<") or s.lower().startswith("before"):
                return "2015 or earlier"
            try:
                dt = pd.to_datetime(s)
                q = (dt.month - 1) // 3 + 1
                return f"Q{q}'{str(dt.year)[2:]}"
            except Exception:
                return s[:10]

        def _parse_date(val):
            """Parse a date value, return pd.Timestamp or None."""
            if not val or (isinstance(val, float) and pd.isna(val)):
                return None
            s = str(val).strip()
            if s.startswith("<") or s.lower().startswith("before"):
                return pd.Timestamp("2014-01-01")
            try:
                return pd.to_datetime(s)
            except Exception:
                return None

        def _dep_end(mp_ramp, region):
            """Calculate depreciation end date from MP ramp + region rule."""
            dt = _parse_date(mp_ramp)
            if dt is None:
                return "-"
            years = _DEP_YEARS.get(region, 5)
            end = dt + pd.DateOffset(years=years)
            return f"Q{(end.month-1)//3+1}'{str(end.year)[2:]}"

        def _parse_phase(phase_str):
            """Parse phase string into structured info.

            Examples:
              '1'   -> base=1, suffix='', event='Launch'
              '1O'  -> base=1, suffix='O', event='LTPO Upgrade'
              '2F'  -> base=2, suffix='F', event='Foldable Conversion'
              '1OF' -> base=1, suffix='OF', event='LTPO + Foldable Upgrade'
              '1_1' -> base=1, suffix='_1', event='Sub-phase 1'
              '1_1F'-> base=1, suffix='_1F', event='Sub-phase 1 Foldable'
            """
            s = str(phase_str).strip()
            # Extract base number and suffix
            m = re.match(r'^(\d+)(.*)', s)
            if not m:
                return {"raw": s, "base": 0, "suffix": s, "event": "Unknown"}
            base = int(m.group(1))
            suffix = m.group(2)

            if suffix == "":
                event = "Phase Launch"
            elif suffix == "O":
                event = "LTPO Upgrade"
            elif suffix == "F":
                event = "Form Factor Change"
            elif suffix == "OF":
                event = "LTPO + Form Factor Change"
            elif suffix.startswith("_"):
                sub = suffix.replace("_", "").replace("F", "")
                has_f = "F" in suffix
                event = f"Sub-phase {sub}" + (" Foldable" if has_f else "")
            else:
                event = f"Variant ({suffix})"

            is_upgrade = "O" in suffix and not suffix.startswith("_")
            is_expansion = suffix == "" or suffix.startswith("_") and "F" not in suffix
            return {"raw": s, "base": base, "suffix": suffix, "event": event,
                    "is_upgrade": is_upgrade, "is_expansion": is_expansion}

        def _compute_family_capacity(phases):
            """Compute correct capacity using phase-family grouping.

            Phase suffix rules (applied to same physical capacity tranche):
              - Base number (1,2,3…) = new capacity added
              - O suffix  = LTPO backplane upgrade (same line)
              - F suffix  = Foldable substrate conversion (same line)
              - OF suffix = LTPO + Foldable (same line)
              - _N suffix = Sub-phase within same tranche

            Algorithm:
              1. Group phases by base number → "families"
              2. Within each family, group by MP Ramp date
              3. Take the latest date-group whose total capacity > 0
                 (phases at the same date = capacity split, sum them)
              4. Sum across families for factory total

            Returns dict with all breakdowns.
            """
            from collections import defaultdict

            # Group into families by base number
            families = defaultdict(list)
            for p in phases:
                base = p["parsed"]["base"]
                families[base].append(p)

            # Per-family: pick latest date-group with capacity > 0
            family_results = []
            for base_num in sorted(families.keys()):
                members = families[base_num]
                # Group by mp_ramp date
                date_groups = defaultdict(list)
                for p in members:
                    dt = p["mp_dt"]
                    key = str(dt.date()) if dt else "unknown"
                    date_groups[key].append(p)

                # Sort date keys chronologically, pick latest with capacity
                sorted_dates = sorted(date_groups.keys())
                chosen_group = None
                chosen_date = None
                for dk in reversed(sorted_dates):
                    grp = date_groups[dk]
                    total = sum(p["tft_max_input"] for p in grp)
                    if total > 0:
                        chosen_group = grp
                        chosen_date = dk
                        break
                # If all groups have 0 capacity, take the latest group
                if chosen_group is None:
                    chosen_group = date_groups[sorted_dates[-1]]
                    chosen_date = sorted_dates[-1]

                cap = sum(p["tft_max_input"] for p in chosen_group)
                oled_cap = sum(p["oled_max_input"] for p in chosen_group)
                octa_cap = sum(p["octa_ksheets"] for p in chosen_group)

                # OLED MG equivalent: normalize to TFT sheet size
                # If OLED > TFT for a phase, OLED uses smaller (half-cut) glass
                oled_mg = 0.0
                for p in chosen_group:
                    tft = p["tft_max_input"]
                    oled = p["oled_max_input"]
                    if tft > 0 and oled > 0:
                        ratio = max(1, round(oled / tft))
                        oled_mg += oled / ratio
                    elif oled > 0:
                        oled_mg += oled  # No TFT reference, use raw

                family_results.append({
                    "base": base_num,
                    "members": [p["phase_raw"] for p in members],
                    "latest_group": chosen_group,
                    "latest_date": chosen_date,
                    "tft_capacity": cap,
                    "oled_capacity": oled_cap,
                    "oled_mg_equiv": oled_mg,
                    "octa_capacity": octa_cap,
                })

            # Aggregate across families
            total_tft = sum(f["tft_capacity"] for f in family_results)
            total_oled = sum(f["oled_capacity"] for f in family_results)
            total_oled_mg = sum(f["oled_mg_equiv"] for f in family_results)
            total_octa = sum(f["octa_capacity"] for f in family_results)

            # Old (incorrect) sum for comparison
            old_sum = sum(p["tft_max_input"] for p in phases)

            # Effective capacity = min of process stages (in MG-equivalent)
            process_stages = {}
            if total_tft > 0:
                process_stages["TFT"] = total_tft
            if total_oled_mg > 0:
                process_stages["OLED"] = total_oled_mg
            if total_octa > 0:
                process_stages["OCT"] = total_octa
            if process_stages:
                bottleneck_stage = min(process_stages, key=process_stages.get)
                effective_cap = process_stages[bottleneck_stage]
                has_bottleneck = (len(process_stages) > 1 and
                                  effective_cap < max(process_stages.values()) * 0.9)
            else:
                bottleneck_stage = None
                effective_cap = total_tft
                has_bottleneck = False

            # Technology split by backplane (from latest config per family)
            tech_split = defaultdict(float)
            for f in family_results:
                for p in f["latest_group"]:
                    tech_split[p["backplane"]] += p["tft_max_input"]

            # Form factor breakdown (from latest config per family)
            standard_rigid_cap = 0.0
            thin_profile_cap = 0.0
            foldable_cap = 0.0
            for f in family_results:
                for p in f["latest_group"]:
                    c = p["tft_max_input"]
                    sub = str(p.get("substrate") or "").strip()
                    enc = str(p.get("encapsulation") or "").strip()
                    if sub == "Rigid" and enc != "TFE":
                        standard_rigid_cap += c
                    elif sub == "Rigid" and enc == "TFE":
                        thin_profile_cap += c
                    elif sub in ("Flexible", "Foldable", "Hybrid", "Rigid/Flexible"):
                        foldable_cap += c
                    else:
                        standard_rigid_cap += c  # Unknown → default

            glass_sub = standard_rigid_cap + thin_profile_cap
            pi_sub = foldable_cap

            # Application breakdown (from latest config per family)
            app_split = defaultdict(float)
            for f in family_results:
                for p in f["latest_group"]:
                    app = str(p.get("main_application") or "-").strip()
                    if not app or app == "-":
                        app = str(p.get("application") or "Unknown").strip()
                    app_split[app] += p["tft_max_input"]

            # 100% LTPO conversion check
            bases_with_o = set()
            for p in phases:
                if "O" in p["parsed"]["suffix"]:
                    bases_with_o.add(p["parsed"]["base"])
            base_only = {p["parsed"]["base"] for p in phases if p["parsed"]["suffix"] == ""}
            all_converted = (len(base_only) > 0 and
                             all(b in bases_with_o for b in base_only))

            return {
                "families": family_results,
                "total_tft": total_tft,
                "total_oled": total_oled,
                "total_oled_mg": total_oled_mg,
                "total_octa": total_octa,
                "old_sum": old_sum,
                "effective_cap": effective_cap,
                "bottleneck_stage": bottleneck_stage,
                "has_bottleneck": has_bottleneck,
                "process_stages": dict(process_stages),
                "tech_split": dict(tech_split),
                "all_converted": all_converted,
                "standard_rigid_cap": standard_rigid_cap,
                "thin_profile_cap": thin_profile_cap,
                "foldable_cap": foldable_cap,
                "glass_sub": glass_sub,
                "pi_sub": pi_sub,
                "app_split": dict(app_split),
            }

        # ── Load ScenarioByFab from Excel (cached) ──
        @st.cache_data(ttl=600)
        def _load_scenario_by_fab():
            """Load ScenarioByFab sheet from CapacityData Excel file."""
            import openpyxl
            src = Path(__file__).parent.parent / "source_data"
            candidates = list(src.glob("*CapSpendReport*CapacityData*"))
            if not candidates:
                return pd.DataFrame()
            fpath = candidates[0]
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            except Exception:
                return pd.DataFrame()
            if "ScenarioByFab" not in wb.sheetnames:
                wb.close()
                return pd.DataFrame()
            ws = wb["ScenarioByFab"]
            # Headers on row 7, data starts row 9
            col_map = {
                3: "region", 4: "manufacturer", 5: "factory1", 6: "location",
                7: "phase", 8: "backplane",
                9: "tft_mg_v", 10: "tft_mg_h", 11: "tft_gen",
                12: "tft_max_input", 13: "octa_ksheets", 14: "octa_mp",
                15: "oled_mg_v", 16: "oled_mg_h", 17: "oled_gen",
                18: "oled_max_input",
                19: "application", 20: "main_application",
                21: "type", 22: "substrate", 23: "depo", 24: "encapsulation",
                25: "eqpt_po", 26: "install", 27: "mp_ramp", 28: "end",
                29: "status", 30: "probability", 31: "client",
                32: "standard_panel",
            }
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=9, max_col=32, values_only=True), start=9):
                if row[2] is None and row[3] is None:
                    continue  # skip empty
                record = {}
                for ci, name in col_map.items():
                    val = row[ci - 1] if ci - 1 < len(row) else None
                    record[name] = val
                rows.append(record)
            wb.close()
            df = pd.DataFrame(rows)
            # Convert numeric columns
            for c in ["tft_max_input", "octa_ksheets", "oled_max_input"]:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
            return df

        scenario_df = _load_scenario_by_fab()
        has_scenario = len(scenario_df) > 0

        # ── Build factory selection list ──
        # Use ScenarioByFab if available (richer data), fallback to DB
        if has_scenario:
            fab_options = (
                scenario_df.groupby(["manufacturer", "factory1"])
                .agg(location=("location", "first"),
                     total_cap=("tft_max_input", "sum"))
                .reset_index()
                .sort_values("total_cap", ascending=False)
            )
            fab_options["label"] = fab_options.apply(
                lambda r: f"{r['manufacturer']} - {r['factory1']} ({r['location'] or ''})", axis=1)
            label_to_key = dict(zip(
                fab_options["label"],
                list(zip(fab_options["manufacturer"], fab_options["factory1"]))
            ))
            default_labels = fab_options["label"].head(3).tolist()
        else:
            all_factories = DatabaseManager.get_factories()
            if len(all_factories) == 0:
                st.info("No factory data available.")
                st.stop()
            fab_options = (
                all_factories.groupby(["manufacturer", "factory_name"])
                .first().reset_index()
            )
            fab_options["label"] = fab_options.apply(
                lambda r: f"{r['manufacturer']} - {r['factory_name']} ({r.get('location') or r.get('region', '')})", axis=1)
            label_to_key = dict(zip(
                fab_options["label"],
                list(zip(fab_options["manufacturer"], fab_options["factory_name"]))
            ))
            default_labels = []

        st.markdown("#### Select Factories to Compare")
        st.caption("Choose 2–4 factories for side-by-side comparison. Default: top 3 by capacity.")

        selected_labels = st.multiselect(
            "Factories",
            options=fab_options["label"].tolist(),
            default=default_labels,
            max_selections=4,
            key="compare_factories_v2",
            label_visibility="collapsed",
        )

        if len(selected_labels) < 2:
            st.info("Select at least 2 factories above to begin comparing.")
        else:
            # ── Gather comprehensive data per factory ──
            compare_data = []
            for label in selected_labels:
                mfr, fname = label_to_key[label]

                # --- Phase data from ScenarioByFab ---
                phases = []
                process_available = False
                if has_scenario:
                    frows = scenario_df[
                        (scenario_df["manufacturer"] == mfr) &
                        (scenario_df["factory1"] == fname)
                    ].copy()
                    if len(frows) > 0:
                        process_available = True
                        region = str(frows.iloc[0].get("region") or "-")
                        location = str(frows.iloc[0].get("location") or "-")
                        tft_gen = str(frows.iloc[0].get("tft_gen") or "-")
                        oled_gen = str(frows.iloc[0].get("oled_gen") or "-")
                        application = str(frows.iloc[0].get("application") or "-")
                        substrate = str(frows.iloc[0].get("substrate") or "-")

                        for _, pr in frows.iterrows():
                            pi = _parse_phase(pr["phase"])
                            mp_dt = _parse_date(pr["mp_ramp"])
                            phases.append({
                                "phase_raw": str(pr["phase"]),
                                "parsed": pi,
                                "backplane": str(pr.get("backplane") or "-"),
                                "tft_max_input": float(pr.get("tft_max_input") or 0),
                                "oled_max_input": float(pr.get("oled_max_input") or 0),
                                "octa_ksheets": float(pr.get("octa_ksheets") or 0),
                                "eqpt_po": _fmt_mp(pr.get("eqpt_po")),
                                "install": _fmt_mp(pr.get("install")),
                                "mp_ramp": _fmt_mp(pr.get("mp_ramp")),
                                "mp_ramp_raw": pr.get("mp_ramp"),
                                "mp_dt": mp_dt,
                                "end": _fmt_mp(pr.get("end")),
                                "dep_end": _dep_end(pr.get("mp_ramp"), region),
                                "status": str(pr.get("status") or "-"),
                                "probability": str(pr.get("probability") or "-"),
                                "client": str(pr.get("client") or "-"),
                                "event": pi["event"],
                                "is_upgrade": pi.get("is_upgrade", False),
                                "is_expansion": pi.get("is_expansion", False),
                                "substrate": str(pr.get("substrate") or "-"),
                                "depo": str(pr.get("depo") or "-"),
                                "encapsulation": str(pr.get("encapsulation") or "-"),
                                "main_application": str(pr.get("main_application") or "-"),
                            })
                        # Sort phases chronologically by MP ramp
                        phases.sort(key=lambda p: p["mp_dt"] or pd.Timestamp("2099-01-01"))

                # --- DB fallback for factory info ---
                db_factory_name = fname
                fdf = DatabaseManager.get_factory_by_name(db_factory_name)
                if fdf is None or len(fdf) == 0:
                    # Try matching via manufacturer
                    all_f = DatabaseManager.get_factories(manufacturer=mfr)
                    match = all_f[all_f["factory_name"] == fname]
                    if len(match) > 0:
                        fdf = match
                    else:
                        continue

                info = fdf.iloc[0]
                if not phases:
                    # No ScenarioByFab data—use DB
                    region = str(info.get("region") or "-")
                    location = str(info.get("location") or "-")
                    tft_gen = str(info.get("generation") or "-")
                    oled_gen = "-"
                    application = str(info.get("application_category") or "-")
                    substrate = str(info.get("substrate") or "-")
                    for _, row in fdf.iterrows():
                        bp = str(row.get("backplane") or "-")
                        phases.append({
                            "phase_raw": "1",
                            "parsed": _parse_phase("1"),
                            "backplane": bp,
                            "tft_max_input": 0, "oled_max_input": 0, "octa_ksheets": 0,
                            "eqpt_po": _fmt_mp(row.get("eqpt_po_year")),
                            "install": _fmt_mp(row.get("install_date")),
                            "mp_ramp": _fmt_mp(row.get("mp_ramp_date")),
                            "mp_ramp_raw": row.get("mp_ramp_date"),
                            "mp_dt": _parse_date(row.get("mp_ramp_date")),
                            "end": "-", "dep_end": _dep_end(row.get("mp_ramp_date"), region),
                            "status": str(row.get("status") or "-"),
                            "probability": str(row.get("probability") or "-"),
                            "client": "-", "event": "Phase Launch",
                            "is_upgrade": False, "is_expansion": True,
                            "substrate": substrate, "depo": "-", "encapsulation": "-",
                            "main_application": application,
                        })
                    phases.sort(key=lambda p: p["mp_dt"] or pd.Timestamp("2099-01-01"))

                # --- Utilization from DB ---
                util = DatabaseManager.get_utilization(factory_name=fname)
                latest_cap, latest_input, latest_util = 0.0, 0.0, 0.0
                if len(util) > 0:
                    util_actual = util[util["actual_input_ksheets"] > 0]
                    ld = util_actual["date"].max() if len(util_actual) > 0 else util["date"].max()
                    latest = util[util["date"] == ld]
                    latest_cap = latest["capacity_ksheets"].sum()
                    latest_input = latest["actual_input_ksheets"].sum()
                    latest_util = (latest_input / latest_cap * 100) if latest_cap > 0 else 0

                # --- Equipment orders from DB ---
                factory_ids = fdf["factory_id"].tolist()
                equip_dfs = [DatabaseManager.get_equipment_orders_for_factory(fid) for fid in factory_ids]
                equip = pd.concat(equip_dfs, ignore_index=True) if equip_dfs else pd.DataFrame()
                if len(equip) > 0:
                    equip = equip.drop_duplicates()
                total_investment = equip["amount_usd"].sum() if len(equip) > 0 and "amount_usd" in equip.columns else 0

                # --- Computed aggregates (using phase-family logic) ---
                # Earliest MP ramp
                ramp_dates = [p["mp_dt"] for p in phases if p["mp_dt"] is not None]
                earliest_ramp_dt = min(ramp_dates) if ramp_dates else None
                earliest_ramp = _fmt_mp(earliest_ramp_dt) if earliest_ramp_dt else "-"
                if any(str(p.get("mp_ramp_raw", "")).lower().startswith("before") or
                       str(p.get("mp_ramp_raw", "")).startswith("<") for p in phases):
                    earliest_ramp = "2015 or earlier"

                # Correct capacity via phase-family grouping
                cap_data = _compute_family_capacity(phases)

                dep_years = _DEP_YEARS.get(region, 5)

                # Depreciation year range
                dep_start_yr = earliest_ramp_dt.year if earliest_ramp_dt else None
                latest_ramp_dts = [p["mp_dt"] for p in phases if p["mp_dt"] is not None]
                latest_ramp_dt = max(latest_ramp_dts) if latest_ramp_dts else None
                dep_end_yr = (latest_ramp_dt.year + dep_years) if latest_ramp_dt else None
                dep_range = (f"{dep_start_yr}-{dep_end_yr}"
                             if dep_start_yr and dep_end_yr else "-")

                compare_data.append({
                    "label": label, "factory_name": fname, "manufacturer": mfr,
                    "location": location, "region": region,
                    "technology": str(info.get("technology") or "-"),
                    "tft_gen": tft_gen, "oled_gen": oled_gen,
                    "application": application, "substrate": substrate,
                    "status": str(info.get("status") or "-").title(),
                    "capacity": latest_cap, "input": latest_input,
                    "utilization": latest_util,
                    "total_investment": total_investment,
                    "earliest_ramp": earliest_ramp,
                    "dep_years": dep_years,
                    "dep_range": dep_range,
                    "phases": phases,
                    # Corrected capacity fields (phase-family logic)
                    "families": cap_data["families"],
                    "tech_split": cap_data["tech_split"],
                    "all_converted": cap_data["all_converted"],
                    "total_tft": cap_data["total_tft"],
                    "total_oled": cap_data["total_oled"],
                    "total_oled_mg": cap_data["total_oled_mg"],
                    "total_octa": cap_data["total_octa"],
                    "old_sum": cap_data["old_sum"],
                    "effective_cap": cap_data["effective_cap"],
                    "bottleneck_stage": cap_data["bottleneck_stage"],
                    "has_bottleneck": cap_data["has_bottleneck"],
                    "process_stages": cap_data["process_stages"],
                    "standard_rigid_cap": cap_data["standard_rigid_cap"],
                    "thin_profile_cap": cap_data["thin_profile_cap"],
                    "foldable_cap": cap_data["foldable_cap"],
                    "glass_sub": cap_data["glass_sub"],
                    "pi_sub": cap_data["pi_sub"],
                    "app_split": cap_data["app_split"],
                    "process_available": process_available,
                    "util_df": util, "equip_df": equip,
                    "factory_df": fdf,
                })

            if len(compare_data) < 2:
                st.warning("Could not load data for enough factories.")
            else:
                n = len(compare_data)

                # ════════════════════════════════════════════
                # COMPARISON CARDS
                # ════════════════════════════════════════════
                card_cols = st.columns(n)
                for i, d in enumerate(compare_data):
                    with card_cols[i]:
                        tech = d["technology"]

                        # Header
                        st.subheader(f"{d['manufacturer']} {d['factory_name']} | {tech}")
                        st.caption(f"{d['location']}, {d['region']}  |  First MP: {d['earliest_ramp']}  |  Depreciation: {d['dep_range']}")

                        # Corrected capacity (phase-family logic)
                        corrected_cap = d["total_tft"]
                        m1, m2 = st.columns(2)
                        with m1:
                            st.metric("Total Capacity", f"{corrected_cap:,.1f}K MG/mo")
                        with m2:
                            st.metric("Utilization", f"{d['utilization']:.1f}%")

                        # Technology Mix
                        ts_parts = []
                        total_ts = sum(d["tech_split"].values())
                        for bp, cap in sorted(d["tech_split"].items(), key=lambda x: -x[1]):
                            if cap > 0:
                                pct = (cap / total_ts * 100) if total_ts > 0 else 0
                                ts_parts.append(f"{bp}: {cap:,.0f}K ({pct:.0f}%)")
                        if d["all_converted"]:
                            ts_parts.append("100% LTPO converted")
                        if ts_parts:
                            st.markdown(f"**Technology:** {' / '.join(ts_parts)}")

                        # Form Factor one-liner
                        ff_parts = []
                        if d["standard_rigid_cap"] > 0:
                            ff_parts.append(f"Standard Rigid: {d['standard_rigid_cap']:,.0f}K")
                        if d["thin_profile_cap"] > 0:
                            ff_parts.append(f"Thin Profile: {d['thin_profile_cap']:,.0f}K")
                        if d["foldable_cap"] > 0:
                            ff_parts.append(f"Foldable: {d['foldable_cap']:,.0f}K")
                        if ff_parts:
                            st.markdown(f"**Form Factor:** {' / '.join(ff_parts)}")

                        # Effective capacity with bottleneck
                        if d["has_bottleneck"]:
                            st.caption(
                                f"Effective Capacity: {d['effective_cap']:,.0f}K "
                                f"(limited by {d['bottleneck_stage']})"
                            )

                        st.divider()

                # ════════════════════════════════════════════
                # EXPANDABLE: Capacity Breakdown (per factory)
                # ════════════════════════════════════════════
                st.divider()
                st.markdown("#### Capacity Breakdown")

                bd_cols = st.columns(n)
                for i, d in enumerate(compare_data):
                    with bd_cols[i]:
                        total = d["total_tft"]
                        with st.expander(f"{d['manufacturer']} {d['factory_name']} — {total:,.0f}K MG/mo", expanded=False):
                            # By Technology (backplane)
                            st.markdown("**Technology Mix**")
                            for bp, cap in sorted(d["tech_split"].items(), key=lambda x: -x[1]):
                                if cap > 0:
                                    pct = (cap / total * 100) if total > 0 else 0
                                    st.markdown(f"- {bp}: **{cap:,.0f}K** ({pct:.0f}%)")
                            if d["all_converted"]:
                                st.caption("100% LTPO (converted from LTPS)")

                            # Form Factor
                            st.markdown("**Form Factor**")
                            if d["standard_rigid_cap"] > 0:
                                pct = (d["standard_rigid_cap"] / total * 100) if total > 0 else 0
                                st.markdown(f"- Standard Rigid: **{d['standard_rigid_cap']:,.0f}K** (glass + glass seal) ({pct:.0f}%)")
                            if d["thin_profile_cap"] > 0:
                                pct = (d["thin_profile_cap"] / total * 100) if total > 0 else 0
                                st.markdown(f"- Thin Profile: **{d['thin_profile_cap']:,.0f}K** (glass + TFE) ({pct:.0f}%)")
                            if d["foldable_cap"] > 0:
                                pct = (d["foldable_cap"] / total * 100) if total > 0 else 0
                                st.markdown(f"- Foldable: **{d['foldable_cap']:,.0f}K** (PI + TFE) ({pct:.0f}%)")
                            if d["standard_rigid_cap"] == 0 and d["thin_profile_cap"] == 0 and d["foldable_cap"] == 0:
                                st.caption("Form factor: N/A")

                            # Process Capacity (TFT/OLED/OCT)
                            st.markdown("**Process Capacity**")
                            if d["process_available"] and (d["total_tft"] > 0 or d["total_oled"] > 0):
                                tft = d["total_tft"]
                                oled_mg = d["total_oled_mg"]
                                oled_raw = d["total_oled"]
                                octa = d["total_octa"]
                                st.markdown(f"- TFT/Backplane Input: **{tft:,.0f}K** MG/mo")
                                if oled_mg != oled_raw and oled_raw > 0:
                                    st.markdown(f"- OLED Encapsulation: **{oled_mg:,.0f}K** MG/mo ({oled_raw:,.0f}K raw)")
                                elif oled_mg > 0:
                                    st.markdown(f"- OLED Encapsulation: **{oled_mg:,.0f}K** MG/mo")
                                if octa > 0:
                                    st.markdown(f"- On-Cell Touch (OCT): **{octa:,.0f}K** MG/mo")
                                else:
                                    st.markdown("- On-Cell Touch (OCT): N/A")

                                # Effective capacity and bottleneck
                                eff = d["effective_cap"]
                                if d["has_bottleneck"]:
                                    st.markdown(
                                        f"- **Effective Capacity: {eff:,.0f}K** "
                                        f"(limited by {d['bottleneck_stage']})"
                                    )
                                elif eff > 0:
                                    st.markdown(f"- Effective Capacity: **{eff:,.0f}K** MG/mo")
                            else:
                                st.caption("Process capacity: N/A")

                            # Application Mix
                            st.markdown("**Application Mix**")
                            if d["app_split"]:
                                for app, cap in sorted(d["app_split"].items(), key=lambda x: -x[1]):
                                    if cap > 0:
                                        pct = (cap / total * 100) if total > 0 else 0
                                        st.markdown(f"- {app}: **{cap:,.0f}K** ({pct:.0f}%)")
                            else:
                                st.caption("Application data: N/A")

                            # Phase family summary
                            st.markdown("**Phase Families**")
                            for fam in d["families"]:
                                members_str = ", ".join(fam["members"])
                                latest_phases = fam["latest_group"]
                                latest_bp = ", ".join(sorted(set(p["backplane"] for p in latest_phases)))
                                latest_sub = ", ".join(sorted(set(str(p.get("substrate", "-")) for p in latest_phases)))
                                latest_enc = ", ".join(sorted(set(str(p.get("encapsulation", "-")) for p in latest_phases)))
                                # Form factor label
                                ff = "Standard Rigid"
                                if latest_enc == "TFE" and latest_sub in ("Flexible", "Foldable", "Hybrid"):
                                    ff = "Foldable"
                                elif latest_enc == "TFE" and latest_sub == "Rigid":
                                    ff = "Thin Profile"
                                st.markdown(
                                    f"- Family {fam['base']} [{members_str}] → "
                                    f"**{fam['tft_capacity']:,.0f}K** ({latest_bp}, {ff})"
                                )

                # ════════════════════════════════════════════
                # EXPANDABLE: Investment & Capacity Timeline
                # ════════════════════════════════════════════
                st.divider()
                st.markdown("#### Investment & Capacity Timeline")

                # Investment data from historical CapSpendReport files to be added
                tl_cols = st.columns(n)
                for i, d in enumerate(compare_data):
                    with tl_cols[i]:
                        num_fam = len(d["families"])
                        with st.expander(f"{d['manufacturer']} {d['factory_name']} — {num_fam} families, {len(d['phases'])} phases", expanded=False):
                            # Group phases by family for display
                            from collections import defaultdict as _dd
                            fam_phases = _dd(list)
                            for p in d["phases"]:
                                fam_phases[p["parsed"]["base"]].append(p)

                            cumulative_cap = 0.0

                            for fam in d["families"]:
                                base = fam["base"]
                                members = fam_phases.get(base, [])
                                # Sort chronologically
                                members_sorted = sorted(members, key=lambda p: p["mp_dt"] or pd.Timestamp("2099-01-01"))

                                st.markdown(f"**Phase {base} Family** (current: {fam['tft_capacity']:,.0f}K MG/mo)")

                                prev_bp = None
                                prev_sub = None
                                prev_enc = None
                                for p in members_sorted:
                                    phase_label = p["phase_raw"]
                                    bp = p["backplane"]
                                    sub = str(p.get("substrate") or "-")
                                    enc = str(p.get("encapsulation") or "-")
                                    tft_cap = p["tft_max_input"]
                                    oled_cap = p["oled_max_input"]
                                    octa_cap = p["octa_ksheets"]
                                    suffix = p["parsed"]["suffix"]

                                    # Form factor label
                                    if sub == "Rigid" and enc != "TFE":
                                        ff = "Standard Rigid"
                                    elif sub == "Rigid" and enc == "TFE":
                                        ff = "Thin Profile"
                                    elif sub in ("Flexible", "Foldable", "Hybrid", "Rigid/Flexible"):
                                        ff = "Foldable"
                                    else:
                                        ff = "Standard Rigid"

                                    # Event classification
                                    if suffix == "":
                                        tag = "NEW CAPACITY"
                                        cap_note = f"+{tft_cap:,.0f}K MG/mo"
                                    elif "O" in suffix and "F" not in suffix:
                                        tag = "TECHNOLOGY UPGRADE"
                                        bp_change = f" ({prev_bp} → {bp})" if prev_bp and prev_bp != bp else ""
                                        cap_note = f"{tft_cap:,.0f}K MG/mo (no net change{bp_change})"
                                    elif "F" in suffix and "O" not in suffix:
                                        tag = "FORM FACTOR CHANGE"
                                        ff_note = f" → {ff}" if prev_sub and prev_sub != sub else ""
                                        cap_note = f"{tft_cap:,.0f}K MG/mo (no net change{ff_note})"
                                    elif "O" in suffix and "F" in suffix:
                                        tag = "TECHNOLOGY + FORM FACTOR CHANGE"
                                        cap_note = f"{tft_cap:,.0f}K MG/mo (no net change)"
                                    else:
                                        tag = "PHASE UPDATE"
                                        cap_note = f"{tft_cap:,.0f}K MG/mo"

                                    # Process capacity line
                                    proc_parts = [f"TFT: {tft_cap:,.0f}K"]
                                    if oled_cap > 0:
                                        proc_parts.append(f"OLED: {oled_cap:,.0f}K")
                                    if octa_cap > 0:
                                        proc_parts.append(f"OCT: {octa_cap:,.0f}K")
                                    proc_str = ", ".join(proc_parts)

                                    # Investment placeholder
                                    inv_key = (d["manufacturer"], d["factory_name"], phase_label)
                                    inv_amt = _PHASE_INVESTMENT.get(inv_key, None)
                                    inv_str = f"${inv_amt/1e6:,.0f}M" if inv_amt else "TBD"

                                    st.markdown(
                                        f"**{p['mp_ramp']}**: Phase {phase_label} — {tag}  \n"
                                        f"Backplane: {bp} | Form Factor: {ff}  \n"
                                        f"Substrate: {sub} | Encap: {enc}  \n"
                                        f"Capacity: {cap_note} ({proc_str})  \n"
                                        f"Investment: {inv_str} | Eqpt PO: {p['eqpt_po']} | Install: {p['install']}"
                                    )

                                    prev_bp = bp
                                    prev_sub = sub
                                    prev_enc = enc

                                # Current configuration (latest group)
                                latest = fam["latest_group"]
                                latest_bp = ", ".join(sorted(set(p["backplane"] for p in latest)))
                                latest_sub = ", ".join(sorted(set(str(p.get("substrate", "-")) for p in latest)))
                                # Form factor for current config
                                latest_enc_val = ", ".join(sorted(set(str(p.get("encapsulation", "-")) for p in latest)))
                                curr_ff = "Standard Rigid"
                                if latest_enc_val == "TFE" and any(str(p.get("substrate", "")) in ("Flexible", "Foldable", "Hybrid") for p in latest):
                                    curr_ff = "Foldable"
                                elif latest_enc_val == "TFE" and all(str(p.get("substrate", "")) == "Rigid" for p in latest):
                                    curr_ff = "Thin Profile"
                                st.caption(
                                    f"Current: {latest_bp} | {curr_ff} | "
                                    f"{fam['tft_capacity']:,.0f}K MG/mo"
                                )

                                cumulative_cap += fam["tft_capacity"]
                                st.markdown("---")

                            # Cumulative totals
                            eff_note = ""
                            if d["has_bottleneck"]:
                                eff_note = f" | Effective: {d['effective_cap']:,.0f}K (limited by {d['bottleneck_stage']})"
                            st.markdown(
                                f"**Factory Total: {cumulative_cap:,.0f}K MG/mo{eff_note}**  \n"
                                f"{num_fam} families, {len(d['phases'])} phase rows  \n"
                                f"Investment: "
                                f"**{'${:,.0f}M'.format(d['total_investment']/1e6) if d['total_investment'] > 0 else 'TBD'}**"
                            )

                # ════════════════════════════════════════════
                # UTILIZATION SECTION
                # ════════════════════════════════════════════
                st.divider()
                st.markdown("#### Utilization: Projected vs Actual")

                # Time period filter
                ucol1, ucol2 = st.columns([1, 3])
                with ucol1:
                    util_period = st.selectbox(
                        "Time granularity",
                        ["Monthly", "Quarterly"],
                        key="compare_util_period",
                    )

                # Build chart: capacity (projected) vs actual input
                fig_util = go.Figure()
                avg_utils = []

                for i, d in enumerate(compare_data):
                    util = d["util_df"]
                    if len(util) == 0:
                        avg_utils.append(("-", 0))
                        continue

                    agg = util.groupby("date").agg(
                        capacity=("capacity_ksheets", "sum"),
                        input=("actual_input_ksheets", "sum"),
                    ).reset_index()
                    agg["date"] = pd.to_datetime(agg["date"])

                    if util_period == "Quarterly":
                        agg["period"] = agg["date"].dt.to_period("Q").astype(str)
                        agg = agg.groupby("period").agg(
                            capacity=("capacity", "mean"),
                            input=("input", "mean"),
                        ).reset_index()
                        x_col = "period"
                    else:
                        agg["period"] = agg["date"].dt.strftime("%Y-%m")
                        x_col = "period"

                    agg["util_pct"] = ((agg["input"] / agg["capacity"]) * 100).round(1)
                    agg_prod = agg[agg["input"] > 0]

                    name = f"{d['manufacturer']} {d['factory_name']}"
                    color = colors[i % len(colors)]

                    # Capacity (projected) - dashed line
                    fig_util.add_trace(go.Scatter(
                        x=agg[x_col].tolist(),
                        y=agg["capacity"].tolist(),
                        mode="lines",
                        name=f"{name} Capacity",
                        line=dict(color=color, width=1.5, dash="dash"),
                        hovertemplate=f"{name}<br>%{{x}}<br>Capacity: %{{y:,.1f}}K/mo<extra></extra>",
                        legendgroup=name,
                    ))
                    # Actual input - solid line
                    if len(agg_prod) > 0:
                        fig_util.add_trace(go.Scatter(
                            x=agg_prod[x_col].tolist(),
                            y=agg_prod["input"].tolist(),
                            mode="lines+markers",
                            name=f"{name} Actual",
                            line=dict(color=color, width=2),
                            marker=dict(size=4),
                            hovertemplate=f"{name}<br>%{{x}}<br>Input: %{{y:,.1f}}K/mo<extra></extra>",
                            legendgroup=name,
                        ))

                    # Avg utilization
                    if len(agg_prod) > 0:
                        avg_u = agg_prod["util_pct"].mean()
                        avg_utils.append((name, avg_u))
                    else:
                        avg_utils.append((name, 0))

                apply_chart_theme(fig_util)
                fig_util.update_layout(
                    xaxis_title="Period",
                    yaxis_title="K sheets/mo",
                    height=420,
                    hovermode="x unified",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0, font=dict(size=11)),
                )
                st.plotly_chart(fig_util, use_container_width=True)

                # Avg utilization metrics
                avg_cols = st.columns(n)
                for i, (name, avg_u) in enumerate(avg_utils):
                    with avg_cols[i]:
                        if avg_u > 0:
                            st.metric(f"Avg Util — {name}", f"{avg_u:.1f}%")
                        else:
                            st.metric(f"Avg Util — {name}", "No data")

                # ════════════════════════════════════════════
                # SUMMARY TABLE
                # ════════════════════════════════════════════
                st.divider()
                st.markdown("#### Summary Table")
                summary_rows = []
                for d in compare_data:
                    bp_list = sorted(set(p["backplane"] for p in d["phases"] if p["backplane"] != "-"))
                    summary_rows.append({
                        "Factory": f"{d['manufacturer']} {d['factory_name']}",
                        "Location": f"{d['location']}, {d['region']}",
                        "Gen": d["tft_gen"],
                        "Families": len(d["families"]),
                        "TFT Cap (K)": f"{d['total_tft']:,.0f}",
                        "Effective (K)": f"{d['effective_cap']:,.0f}" if d["effective_cap"] > 0 else "-",
                        "LTPS": f"{d['tech_split'].get('LTPS', 0):,.0f}",
                        "LTPO": f"{d['tech_split'].get('LTPO', 0):,.0f}",
                        "Std Rigid": f"{d['standard_rigid_cap']:,.0f}",
                        "Thin Prof": f"{d['thin_profile_cap']:,.0f}",
                        "Foldable": f"{d['foldable_cap']:,.0f}",
                        "Utilization": f"{d['utilization']:.1f}%",
                        "Depreciation": d["dep_range"],
                    })
                st.dataframe(
                    pd.DataFrame(summary_rows),
                    use_container_width=True,
                    hide_index=True,
                )

